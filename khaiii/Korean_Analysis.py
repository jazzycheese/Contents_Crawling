
#엑셀에서 url가져오기
from openpyxl import load_workbook
import requests
from openpyxl import Workbook
from khaiii import KhaiiiApi

load_wb = load_workbook("/Users/soyeon/dev/khaiii/Chosun/ChosunNewsResult.xlsx", data_only=True)  
load_ws = load_wb['Sheet']

write_wb = Workbook()
write_ws = write_wb.active

api = KhaiiiApi() 
morphs = []
i = 1
j = 

#=========Chekct the coverage=============

get_cells = load_ws['G1':'ADG1']  #Chosun : 'G1': 'ADG1', Kyung : 'B1':'GN1'
for row in get_cells:
    for cell in row:
        #print(cell.value)      
        sentence = str(cell.value)
        for word in api.analyze(sentence):
            for morph in word.morphs:
                morphs.extend((morph.lex, morph.tag))
                if morph.tag == 'NNG':
                    write_ws.cell(row=i, column=j).value = morph.lex
                    #write_ws.append([morph.lex])
                    print(morph.lex)
                    i+=1
        j+=1
        i=1
write_wb.save('/Users/soyeon/dev/NewsAnalysis/WordCount/ChosunCorpus.xlsx')
#from pprint import pprint
#pprint(morphs)