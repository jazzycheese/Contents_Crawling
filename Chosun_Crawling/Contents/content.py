#from bs4 import BeautifulSoup as soup
#from urllib.request import urlopen
import requests
from openpyxl import load_workbook
from openpyxl import Workbook

write_wb = Workbook()
write_ws = write_wb.active

"""
#=========================Get URL=========================
page_url = "file:///Users/soyeon/dev/Contents_Crawling/Chosun_Crawling/Contents/pg3.html"
page = urlopen(page_url)
page_soup = soup(page.read(), "html.parser")

for tag in page_soup.select("td"): 
    print(tag.text) 
    write_ws.append([tag.text])
write_wb.save('Total_Result.xlsx')

#=========================Get Contents=========================
load_wb = load_workbook("Chosun_URL.xlsx", data_only=True)
load_ws = load_wb['Sheet']

get_cells = load_ws['A1':'AAG1']   #'AAG1
for row in get_cells:
    for cell in row:
        #print(cell.value)      
        s = str(cell.value)
        response = requests.get(s)
        html = response.text
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(html, 'html.parser')
        for tag in soup.select("div", class_='article'):
            print(tag.text) 
            write_ws.append([tag.text])
write_wb.save('Chosun_Contents_Result.xlsx')
print('hello world')
"""
#=========================Get Categogy=========================
load_wb = load_workbook("Chosun_URL.xlsx", data_only=True)
load_ws = load_wb['Sheet']

get_cells = load_ws['A1':'AAG1']   #'AAG1
for row in get_cells:
    for cell in row:   
        s = str(cell.value)
        response = requests.get(s)
        html = response.text
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(html, 'html.parser')
        for tag in soup.select("td.list"):
            writer = (tag.text.split("/")[1])
            category = (writer.split(" ")[1])
            print(category) 
            write_ws.append([category])
write_wb.save('Chosun_Category_Result.xlsx')
print('hello world')